"""Excel document converter implementation."""

import os
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Union
from datetime import datetime

import structlog
import openpyxl
from openpyxl.utils import get_column_letter

from morag_core.interfaces.converter import (
    ConversionOptions,
    ConversionError,
)
from morag_core.models.document import Document, DocumentType

from .base import DocumentConverter

logger = structlog.get_logger(__name__)


class ExcelConverter(DocumentConverter):
    """Excel document converter implementation."""

    def __init__(self):
        """Initialize Excel converter."""
        super().__init__()
        self.supported_formats = {"excel", "xlsx", "xls"}

    async def _extract_text(self, file_path: Path, document: Document, options: ConversionOptions) -> Document:
        """Extract text from Excel document.

        Args:
            file_path: Path to Excel file
            document: Document to update
            options: Conversion options

        Returns:
            Updated document

        Raises:
            ConversionError: If text extraction fails
        """
        try:
            # Load workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Extract metadata
            if options.extract_metadata:
                document.metadata.title = os.path.basename(file_path)
                document.metadata.file_type = "excel"
                document.metadata.page_count = len(workbook.sheetnames)
                
                # Try to get properties if available
                try:
                    props = workbook.properties
                    if props.title:
                        document.metadata.title = props.title
                    if props.creator:
                        document.metadata.author = props.creator
                    if props.created:
                        document.metadata.creation_date = props.created.isoformat()
                    if props.modified:
                        document.metadata.modification_date = props.modified.isoformat()
                except Exception as e:
                    logger.warning("Failed to extract Excel properties", error=str(e))
            
            # Process based on chunking strategy
            if options.chunking_strategy == "sheet":
                # Process each sheet as a separate chunk
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet_text = self._extract_sheet_text(sheet)
                    
                    # Add sheet as chunk
                    document.add_chunk(
                        content=sheet_text,
                        section=sheet_name,
                    )
            else:
                # Process all sheets as a single document
                all_text = []
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    sheet_text = self._extract_sheet_text(sheet)
                    all_text.append(f"Sheet: {sheet_name}\n\n{sheet_text}")
                
                # Join all text
                document.raw_text = "\n\n".join(all_text)
                
                # Apply default chunking strategy
                await self._chunk_document(document, options)
            
            # Estimate word count
            document.metadata.word_count = len(document.raw_text.split())
            
            return document
            
        except Exception as e:
            logger.error(
                "Excel document extraction failed",
                error=str(e),
                error_type=e.__class__.__name__,
                file_path=str(file_path),
            )
            raise ConversionError(f"Failed to extract text from Excel document: {str(e)}")
    
    def _extract_sheet_text(self, sheet) -> str:
        """Extract text from Excel sheet.

        Args:
            sheet: Excel worksheet

        Returns:
            Extracted text
        """
        rows = []
        max_column = sheet.max_column
        max_row = sheet.max_row
        
        # Get column headers
        headers = []
        for col in range(1, max_column + 1):
            cell_value = sheet.cell(row=1, column=col).value
            if cell_value is not None:
                headers.append(str(cell_value))
            else:
                headers.append(get_column_letter(col))
        
        # Add headers row
        if headers:
            rows.append(" | ".join(headers))
            rows.append("-" * (sum(len(h) for h in headers) + 3 * (len(headers) - 1)))
        
        # Process data rows
        for row in range(2, max_row + 1):
            row_values = []
            for col in range(1, max_column + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is not None:
                    # Format dates and times
                    if isinstance(cell_value, datetime):
                        cell_value = cell_value.isoformat()
                    row_values.append(str(cell_value))
                else:
                    row_values.append("")
            
            # Only add row if it has at least one non-empty value
            if any(val for val in row_values):
                rows.append(" | ".join(row_values))
        
        return "\n".join(rows)